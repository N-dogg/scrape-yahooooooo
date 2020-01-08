import requests
import pandas as pd
import numpy as np
from bs4 import BeautifulSoup
import re
from openpyxl import Workbook, load_workbook
from dateutil.parser import parse

def is_date(string, fuzzy=False):
    """
    Return whether the string can be interpreted as a date.
    """
    try:
        parse(string, fuzzy=fuzzy)
        return True
    except ValueError:
        return False

#Import stock codes and set up URl, statement and filename variables
codes = pd.read_csv('codes.csv')
url = "https://au.finance.yahoo.com/quote/{0}.AX/{1}?p={0}.AX"
statements = ["financials", "balance-sheet", "cash-flow"]
filename = "{0}.xlsx"

if __name__ == '__main__':
    #Iterate
    for i in codes['Codes']:
        wb = openpyxl.Workbook()
        wb.save(filename.format(i))

    for i in codes['Codes']:
        book = load_workbook(filename.format(i))
        writer = pd.ExcelWriter(filename.format(i), engine = 'openpyxl')
        writer.book = book
        for y in statements:
            line_item = []
            years = []
            data = []
            page = requests.get(url.format(i, y))
            print(url.format(i,y))
            soup = BeautifulSoup(page.content, 'html.parser')
            blob_header = soup.find_all(['span'], class_ = "Va(m)")
            for x in range(len(blob_header)):
                line_item.append(blob_header[x].get_text())
            blob_year = soup.find_all(['div'], attrs={"class": "D(ib)"})
            for z in range(len(blob_year)):
                if is_date(blob_year[z].get_text()) or blob_year[z].get_text()=="ttm":
                    years.append((blob_year[z].get_text()))
            blob_data = soup.find_all(['div'], attrs={"data-test": "fin-col"})
            for w in range(len(blob_data)):
                data.append(blob_data[w].get_text())
            data = pd.DataFrame(np.reshape(data, (len(line_item), len(years))))
            data.index, data.columns = line_item, years
            data.to_excel(writer, y)
            writer.save()
